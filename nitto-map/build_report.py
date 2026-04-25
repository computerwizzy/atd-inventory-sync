import openpyxl, csv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load NTU MAP
wb_in = openpyxl.load_workbook(r'c:\Users\DELL-i7\Downloads\wheel1_not_in _store\NTU MAP Schedule 4.1.2026.xlsx', data_only=True)
ws_in = wb_in.active
ntu = {}
for i, row in enumerate(ws_in.iter_rows(values_only=True)):
    if i < 3:
        continue
    if row[1] and str(row[1]).strip().isdigit():
        code = str(row[1]).strip()
        ntu[code] = {
            'pattern': row[0],
            'description': row[2],
            'map_mar': row[3],
            'map_apr': row[4],
            'change': row[6]
        }

# Load WP feed
wp = {}
with open(r'c:\Users\DELL-i7\Downloads\wheel1_not_in _store\tireInvPriceData.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row['Brand'].strip().lower() == 'nitto':
            pn = row['PartNumber'].strip().lstrip('N').replace('-', '')
            wp[pn] = {
                'wp_pn': row['PartNumber'],
                'wp_desc': row['PartDescription'],
                'msrp': float(row['MSRP_USD'] or 0),
                'map_wp': float(row['MAP_USD'] or 0),
                'qty': int(row['TotalQOH'] or 0)
            }

matched = []
ntu_only = []
for code, n in ntu.items():
    if code in wp:
        w = wp[code]
        apr_map = n['map_apr']
        wp_map = w['map_wp']
        diff = round(wp_map - apr_map, 2) if apr_map else None
        if wp_map == apr_map:
            status = 'OK'
        elif wp_map == 0:
            status = 'MAP MISSING IN WP'
        else:
            status = 'MISMATCH'
        matched.append((code, n['pattern'], n['description'], n['map_mar'], apr_map, wp_map, w['msrp'], diff, w['qty'], n['change'], status))
    else:
        ntu_only.append((code, n['pattern'], n['description'], n['map_apr']))

mismatches = [r for r in matched if r[10] != 'OK']
correct = [r for r in matched if r[10] == 'OK']

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
normal_font = Font(name='Calibri', size=11)
dark_fill   = PatternFill('solid', fgColor='1F4E79')
blue_fill   = PatternFill('solid', fgColor='2E75B6')
red_fill    = PatternFill('solid', fgColor='C00000')
green_fill  = PatternFill('solid', fgColor='E2EFDA')
yellow_fill = PatternFill('solid', fgColor='FFF2CC')
light_red   = PatternFill('solid', fgColor='FFDAD9')
light_blue  = PatternFill('solid', fgColor='DEEAF1')

wb = openpyxl.Workbook()

# ── SUMMARY ──────────────────────────────────────────────────────────────────
ws_sum = wb.active
ws_sum.title = 'Summary'

ws_sum.merge_cells('A1:D1')
ws_sum['A1'] = 'Nitto MAP Price Analysis  |  WheelPros Feed vs NTU Schedule (April 1, 2026)'
ws_sum['A1'].font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
ws_sum['A1'].fill = dark_fill
ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_sum.row_dimensions[1].height = 30

for col, label in [('A', 'Metric'), ('B', 'Count')]:
    c = ws_sum[f'{col}3']
    c.value = label
    c.font = header_font
    c.fill = blue_fill
    c.alignment = Alignment(horizontal='center')

summary_rows = [
    ('Total NTU SKUs (April Schedule)', 820, light_blue),
    ('Matched in WheelPros Feed', len(matched), light_blue),
    ('Not in WheelPros Feed', len(ntu_only), light_blue),
    ('MAP Correct in WP', len(correct), green_fill),
    ('MAP Missing in WP (set to $0)', sum(1 for r in matched if r[10] == 'MAP MISSING IN WP'), yellow_fill),
    ('MAP Mismatch (wrong value)', sum(1 for r in matched if r[10] == 'MISMATCH'), light_red),
]

for i, (label, val, fill) in enumerate(summary_rows, start=4):
    ws_sum[f'A{i}'] = label
    ws_sum[f'B{i}'] = val
    ws_sum[f'A{i}'].font = normal_font
    ws_sum[f'B{i}'].font = Font(name='Calibri', bold=True, size=11)
    ws_sum[f'A{i}'].fill = fill
    ws_sum[f'B{i}'].fill = fill
    ws_sum[f'B{i}'].alignment = Alignment(horizontal='center')

ws_sum.column_dimensions['A'].width = 42
ws_sum.column_dimensions['B'].width = 15

# ── ISSUES ────────────────────────────────────────────────────────────────────
ws_iss = wb.create_sheet('Issues - MAP Discrepancies')
headers = ['Part #', 'Pattern', 'NTU Description', 'NTU MAP Mar', 'NTU MAP Apr', 'WP MAP', 'WP MSRP', 'Difference', 'WP QTY', 'NTU Change', 'Status']
widths  = [10, 10, 44, 13, 13, 12, 12, 12, 10, 16, 24]

for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws_iss.cell(1, ci, h)
    c.font = header_font
    c.fill = red_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws_iss.column_dimensions[get_column_letter(ci)].width = w
ws_iss.row_dimensions[1].height = 30

for ri, r in enumerate(mismatches, 2):
    fill = yellow_fill if r[10] == 'MAP MISSING IN WP' else light_red
    vals = [r[0], r[1], r[2],
            f'${r[3]}', f'${r[4]}', f'${r[5]}', f'${r[6]}',
            f'${r[7]}' if r[7] is not None else '-',
            r[8], r[9], r[10]]
    for ci, v in enumerate(vals, 1):
        c = ws_iss.cell(ri, ci, v)
        c.font = normal_font
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

ws_iss.freeze_panes = 'A2'

# ── NOT IN WP ─────────────────────────────────────────────────────────────────
ws_miss = wb.create_sheet('Not in WP Feed')
headers2 = ['Part #', 'Pattern', 'NTU Description', 'NTU MAP Apr']
widths2  = [10, 10, 48, 14]
for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
    c = ws_miss.cell(1, ci, h)
    c.font = header_font
    c.fill = blue_fill
    c.alignment = Alignment(horizontal='center')
    ws_miss.column_dimensions[get_column_letter(ci)].width = w

for ri, r in enumerate(ntu_only, 2):
    for ci, v in enumerate([r[0], r[1], r[2], f'${r[3]}'], 1):
        c = ws_miss.cell(ri, ci, v)
        c.font = normal_font
        c.fill = light_blue
        c.alignment = Alignment(horizontal='center')

ws_miss.freeze_panes = 'A2'

# ── ALL MATCHED ───────────────────────────────────────────────────────────────
ws_all = wb.create_sheet('All Matched SKUs')
headers3 = ['Part #', 'Pattern', 'NTU Description', 'NTU MAP Mar', 'NTU MAP Apr', 'WP MAP', 'WP MSRP', 'Difference', 'WP QTY', 'NTU Change', 'Status']
widths3  = [10, 10, 44, 13, 13, 12, 12, 12, 10, 16, 24]
for ci, (h, w) in enumerate(zip(headers3, widths3), 1):
    c = ws_all.cell(1, ci, h)
    c.font = header_font
    c.fill = dark_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws_all.column_dimensions[get_column_letter(ci)].width = w
ws_all.row_dimensions[1].height = 30

for ri, r in enumerate(matched, 2):
    if r[10] == 'OK':
        fill = green_fill
    elif r[10] == 'MAP MISSING IN WP':
        fill = yellow_fill
    else:
        fill = light_red
    vals = [r[0], r[1], r[2],
            f'${r[3]}', f'${r[4]}', f'${r[5]}', f'${r[6]}',
            f'${r[7]}' if r[7] is not None else '-',
            r[8], r[9], r[10]]
    for ci, v in enumerate(vals, 1):
        c = ws_all.cell(ri, ci, v)
        c.font = normal_font
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

ws_all.freeze_panes = 'A2'

out = r'c:\Users\DELL-i7\Downloads\wheel1_not_in _store\Nitto_MAP_Analysis.xlsx'
wb.save(out)
print(f'Done. Saved to: {out}')
print(f'Total NTU SKUs: 820 | Matched: {len(matched)} | Issues: {len(mismatches)} | Not in WP: {len(ntu_only)}')
